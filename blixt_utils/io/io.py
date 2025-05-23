import matplotlib.pyplot as plt
import pandas as pd
import xarray as xr
import os
import re
import numpy as np
from datetime import datetime
from openpyxl import load_workbook, Workbook
from copy import deepcopy
import getpass
import socket
import logging
from scipy.interpolate import interp1d

import segyio

# If wavelet_plot is used, I get an 'ImportError' "cannot import name 'wavelet_plot' from partially initialized
# module 'blixt_utils.plotting.helpers' (most likely due to a circular import)"
# from blixt_utils.plotting.helpers import wavelet_plot

logger = logging.getLogger(__name__)


def read_segy(f, lag=0, twod=False, byte_il=189, byte_xl=193):
    """
    read_segy (C) aadm 2018 // using Statoil's segyio
    https://nbviewer.jupyter.org/github/aadm/geophysical_notes/blob/master/playing_with_seismic.ipynb

    Slightly modified and upgraded by Erik Mårten Blixt 2020-08-19
    """
    from blixt_utils.utils import print_info
    if twod:
        with segyio.open(f, 'r', ignore_geometry=True) as segyfile:
            sr = segyio.tools.dt(segyfile)/1e3
            nsamples = segyfile.samples.size
            twt = segyfile.samples
            ntraces = segyfile.tracecount
            data = segyfile.trace.raw[:]
            header = segyio.tools.wrap(segyfile.text[0])
    else:
        with segyio.open(f, iline=byte_il, xline=byte_xl) as segyfile:
            sr = segyio.tools.dt(segyfile)/1e3
            nsamples = segyfile.samples.size
            twt = segyfile.samples
            ntraces = segyfile.tracecount
            data = segyio.tools.cube(segyfile)
            header = segyio.tools.wrap(segyfile.text[0])
            inlines = segyfile.ilines
            crosslines = segyfile.xlines
    size_mb = data.nbytes/1024**2
    info_txt = '[read_segy] reading {}\n'.format(f)
    info_txt += '[read_segy] number of traces: {0}, samples: {1}, sample rate: {2} s\n'.format(ntraces, nsamples, sr)
    info_txt += '[read_segy] first, last sample twt: {0}, {1} s\n'.format(twt[0], twt[-1])
    info_txt += '[read_segy] size: {:.2f} Mb ({:.2f} Gb)'.format(size_mb, size_mb/1024)
    print_info(info_txt, 'info', logger)
    if not twod:
        info_txt = '[read_segy] inlines: {:.0f}, min={:.0f}, max={:.0f}\n'.format(
            inlines.size, inlines.min(), inlines.max())
        info_txt += '[read_segy] crosslines: {:.0f}, min={:.0f}, max={:.0f}'.format(
            crosslines.size, crosslines.min(), crosslines.max())
        print_info(info_txt, 'info', logger)
        return xr.DataArray(data, dims=['INLINE', 'XLINE', 'TWT'], coords=[inlines, crosslines, twt]), \
               nsamples, sr, twt, ntraces, header, inlines, crosslines
    else:
        return xr.DataArray(data, dims= ['TRACES', 'TWT'], coords=[np.arange(ntraces), twt]), \
               nsamples, sr, twt, ntraces, header, None, None


def read_petrel_wavelet(filename,
                        normalize=True,
                        convert_to_zero_phase=False,
                        resample_to=None,
                        verbose=False,
                        return_dict=False,
                        **kwargs):
    """
    Reads a wavelet exported from Petrel using the ASCII format
    Returns the header (dictionary), time [s] and the wavelet
    :param filename:
        str
        filename with full path to wavelet file in Petrel ASCII format
    :param normalize:
        bool
        if set to True, the wavelet values are normalized so that the maximum value is one
    :param convert_to_zero_phase:
        bool
        if set to True, the wavelet is converted to zero phase so that its maximum occurs at time zero
    :param resample_to:
        float
        set this to the desired sample rate of the wavelet [s].
    :param verbose:
        bool
        If True, a plot of the wavelet is generated
    """
    name_identifier = 'WAVELET-NAME'
    sample_rate_identifier = 'SAMPLE-RATE'
    name = 'unknown'
    sample_rate = None
    zero_phased = False
    resampled = False
    scale_factor = 1.

    # create empty data containers
    header, time, wavelet = {}, [], []

    with open(filename) as f:
        all_lines = f.readlines()

    inside_header = True
    for line in all_lines:
        if line[:3] == 'EOH':
            inside_header = False
            continue
        if line[:3] == 'EOD':
            break
        if inside_header:
            if name_identifier in line:
                name = line.replace(name_identifier, '').replace('\n', '').strip()
            if sample_rate_identifier in line:
                sample_rate = line.replace(sample_rate_identifier, '').replace('\n', '').strip()
        if not inside_header:
            data = line.replace('\n', '').strip().split()
            if len(data) == 2:
                time.append(float(data[0]) / 1000.)  # ms to s
                wavelet.append(float(data[1]))

    header['Name'] = name
    header['Sample rate'] = float(sample_rate)
    header['Original filename'] = filename
    header['Normalized'] = normalize
    header['Scale factor'] = scale_factor
    header['Converted to zero phase'] = zero_phased
    header['Time shift'] = 0.
    header['Resampled'] = resampled

    wavelet = np.array(wavelet)
    time = np.array(time)
    if normalize:
        scale_factor = max(wavelet)
        wavelet = wavelet / scale_factor
        header['Normalized'] = normalize
        header['Scale factor'] = scale_factor
    if convert_to_zero_phase:
        time_at_max = time[np.argmax(wavelet)]
        time = time - time_at_max
        header['Converted to zero phase'] = True
        header['Time shift'] = -1. * time_at_max

    if resample_to is not None:
        if not isinstance(resample_to, float):
            raise IOError('Sample rate must be a floating number in seconds')
        if resample_to == float(sample_rate):
            print('Desired sample rate {} matches original sample rate {}. No resampling done'.format(
                resample_to, sample_rate
            ))
        else:
            f = interp1d(time, wavelet, kind='cubic')
            #new_time = np.arange(time[0], time[-1] + resample_to, resample_to)
            new_time = np.arange(time[0], time[-1], resample_to)
            new_wavelet = f(new_time)
            time = new_time
            wavelet = new_wavelet
            resampled = True
            header['Sample rate'] = resample_to
            header['Resampled'] = resampled

    if np.mod(len(time), 2) != 0:  # odd number of elements
        time = time[:-1]
        wavelet = wavelet[:-1]

    # if verbose:
    #     wavelet_plot(None, time, wavelet, header)

    if return_dict:
        return {'wavelet': wavelet, 'time': time, 'header': header}
    else:
        return header, time, wavelet


def read_checkshot_or_wellpath(project_table_name, well_name, sheet_name, verbose=False):
    """
    Tries to read well path (deviation) OR checkshot data from an ASCII file in different formats and returns a
    dictionary with the
    different columns as key: value pairs.
    AND a second dictionary with information about the data that's been added
    :param project_table_name:
        str
        full path name to xlsx project table
    :param well_name:
        str
        Given well name
    :param sheet_name:
        str
        Name of the sheet in the xlsx project table from where to fetch the file information
        Either "Checkshots" or "Well paths"
    :return:
        dict, dict
        A dictionary with data, and one with information
    """
    from blixt_utils.utils import print_info
    tmp = project_files_info(project_table_name, sheet_name)
    if well_name not in tmp:
        warn_txt = 'Well {} not listed in {} sheet of {}'.format(well_name, sheet_name, project_table_name)
        print_info(warn_txt, 'warning', logger)
        return None, None

    kwargs = tmp[well_name]

    if sheet_name == 'Well paths':
        path_type = kwargs.pop('type')
        if path_type == 'Vertical':
            return None, project_files_info(project_table_name, sheet_name)[well_name]

    filename = kwargs.pop('filename')
    if filename is None:
        info_txt = 'No filename specified for {}, return None'.format(well_name)
        print_info(info_txt, 'info', logger)
        return None
    if not os.path.isfile(filename):
        # try adding the default working directory to the file path
        dir_list = project_table_name.split(os.path.sep)
        working_dir = os.path.sep.join(dir_list[:-2])
        filename = os.path.join(working_dir, filename)

    file_format = kwargs.pop('file format')
    if file_format is None:
        raise IOError('Well path or checkshot file format not specified')
    if file_format.lower() not in ['well trace from petrel', 'general ascii', 'original survey points',
                                   'petrel checkshot']:
        raise IOError('Unknown file format {}'.format(file_format))

    if file_format.lower() == 'well trace from petrel':
        data = read_well_trace_from_petrel(filename)
    elif file_format.lower() == 'original survey points':
        data = read_original_survey_points(filename)
    elif file_format.lower() == 'petrel checkshot':
        data = read_petrel_checkshots(filename)[well_name]
    elif file_format.lower() == 'general ascii':
        data = read_general_ascii(filename, sheet_name, verbose=verbose, **kwargs)
    else:
        raise IOError('Missing valid file format specifier')

    # convert from lists to arrays
    if data is not None:
        for key in list(data.keys()):
            data[key] = np.array(data[key])

    # For checkshots, convert time from ms to s
    if sheet_name == "Checkshots":
        if "time units" not in kwargs:
            raise IOError("Units of time not provided in {} sheet of {}".format(sheet_name, project_table_name))
        if kwargs["time units"] == "ms":
            data["TWT"] = data["TWT"] / 1000.
    return data, project_files_info(project_table_name, sheet_name)[well_name]


def read_well_trace_from_petrel(filename):
    """
    Reads and returns the well path information from a well trace file from Petrel
    :param filename:
    :return:
        dict
        Dictionary with the MD, TVD and INCLINATION data
    """
    keys = None
    data = None
    with open(filename, 'r') as f:
        for line in f.readlines():
            if line[0] == '#':
                continue
            else:
                if 'MD' in line[:20]:
                    line = line.replace('INCL', 'INC')  # to match blixt_rp add_well_path() function
                    keys = line.split()
                    data = {x: [] for x in keys}
                else:
                    i = 0
                    this_line_of_data = line.split()
                    for key in keys:
                        data[key].append(float(this_line_of_data[i]))
                        i += 1
    return data


def read_original_survey_points(filename):
    """
    Reads and returns the well path information from an "original survey points" file
    :param filename:
    :return:
        dict
        Dictionary with the MD, TVD and INCLINATION data
    """
    data = None
    keys = None
    with open(filename, 'r') as f:
        for line in f.readlines():
            if line[:4] == 'MD  ':
                line = line.replace('Inc', 'INC')
                line = line.replace('TVD(RKB)', 'TVD')
                keys = line.split()
                data = {x: [] for x in keys}
            elif line[0].isnumeric():
                this_line_of_data = _split(line, 'space')
                j = 0
                for key in keys:
                    data[key].append(this_line_of_data[j])
                    j += 1

    return data


def read_general_ascii(filename, data_type, **kwargs):
    """
    Uses the information in kwargs to read an ASCII file with data stored in a "general" format
    :param filename:
        str
        full path name of ASCII file with data
    :param data_type:
        str
        Either "Checkshots" or "Well paths"
    :param kwargs:
        dict
        Dictionary with information about how to interpret the general ascii file
        Typically returned from the project_files_info() function
    :return:
        dict
        Dictionary containing the data in the ascii file with each column of data sorted under a given key
    """
    def contain_strings(l:list):
        return not all([not isinstance(_x, str) for _x in l])

    def return_valid_line(line_string, sep, _num_cols=None):
        # Skip empty lines!
        if len(line_string.strip().replace('\t', '')) == 0:
            return False

        _data = _split(line_string, sep)

        # skip lines shorter than minimum columns
        if _num_cols is not None:
            if len(_data) != _num_cols:
                return False

        # Try to convert all elements of data list to floats
        _data = [my_float(_x) for _x in _data]

        # # skip lines that contain strings
        # if contain_strings(_data):
        #     return False

        return _data

    data_line = int(kwargs.pop('data begins on line'))
    separator = kwargs.pop('separator', 'space')
    separator = separator.replace('"', '')
    md_column = int(kwargs.pop('md column'))
    verbose = kwargs.pop('verbose', False)
    if data_type == "Well paths":
        tvd_column = int(kwargs.pop('tvd column'))
        inclination_column = int(kwargs.pop('inclination column'))
        if verbose:
            print('File: {}'.format(os.path.basename(filename)))
            print(' MD column: {}'.format(md_column))
            print(' TVD column: {}'.format(tvd_column))
            print(' INC column: {}'.format(inclination_column))
        data = {x: [] for x in ['TVD', 'MD', 'INC']}
        with open(filename, 'r') as f:
            i = 0
            num_cols = None
            for line in f.readlines():
                if i >= data_line - 1:
                    this_line_of_data = return_valid_line(line, separator, num_cols)
                    if not this_line_of_data:
                        continue
                    if i == data_line - 1:
                        num_cols = len(this_line_of_data)
                    data['MD'].append(this_line_of_data[md_column - 1])
                    data['TVD'].append(this_line_of_data[tvd_column - 1])
                    data['INC'].append(this_line_of_data[inclination_column - 1])
                i += 1
        return data
    elif data_type == 'Checkshots':
        sign_multiplier = None
        time_column = None
        owt_column = kwargs.pop('owt column')
        twt_column = kwargs.pop('twt column')
        if twt_column is None:
            if owt_column is None:
                raise IOError('Column number for both one-, and two-, way time are lacking')
            else:
                sign_multiplier = 2.
                time_column = int(owt_column)
        else:
            sign_multiplier = 1.
            time_column = int(twt_column)
        data = {x: [] for x in ['MD', 'TWT']}
        with open(filename, 'r') as f:
            i = 0
            num_cols = None
            for line in f.readlines():
                if i >= data_line - 1:
                    this_line_of_data = return_valid_line(line, separator, num_cols)
                    if not this_line_of_data:
                        continue
                    if i == data_line - 1:
                        num_cols = len(this_line_of_data)
                    data['MD'].append(this_line_of_data[md_column - 1])
                    data['TWT'].append(this_line_of_data[time_column - 1] * sign_multiplier)
                i += 1
        return data
    else:
        raise IOError('Unknown data type: {}'.format(data_type))


def read_general_ascii_GENERAL(
        filename: str,
        separator: str,
        data_begins_on_row: int,
        var_names=None,
        var_columns=None,
        var_units=None,
        encoding=None
) -> (dict, dict):
    """
    General ASCII file reader that isn't specific to any type of data.
    It requires that the data are arranged as rows and columns,
    where each column is a data parameter

    :param filename:
    :param separator:
        'space', 'tab', ',', ';', ...
    :param data_begins_on_row:
        int
        Pythonic row number (starting at 0) of the first line of data
    :param var_names:
        int or list
        Either (pythonic) row number of where the variable names can be read in the file (NOTE: must have
        the same number of items and separator as the data for it to work)
        OR
        list of variables names. Its length must match the number of data columns or length of var_columns

        If None: generic variable names are created automatically
    :param var_columns:
        list
        list of integers that indicate the column numbers (pythonic) to load for each variable listed in var_names
        if var_names is an integer, all variables are loaded and this parameter is ignored
    :param var_units:
        int or list
        Either (pythonic) row number of where the units can be read in the file (NOTE: must have
        the same number of items and separator as the data for it to work)
        OR
        list of units. Its length must match the number of data columns

        If None: empty units are assigned each variable
    :param encoding:
        str
        Name of encoding to use when reading the data
        'utf-8-sig' seems useful

    :return:
        data, units
        data: dictionary with variable names as key, and its data (list) as values
        units: dictionary with variable names as key, and its unit (str) as values
    """
    # First make sure we can extract one row of data, and assess the number of variables:
    n = None
    with open(filename, 'r', encoding=encoding) as f:
        for i, line in enumerate(f.readlines()):
            if i == data_begins_on_row:
                n = len(_return_valid_line(line, separator))
                break
    if n is None:
        return None, None

    # Determine variable names
    if var_names is None:
        names = ['var_{}'.format(i+1) for i in np.arange(n)]
    elif isinstance(var_names, int):
        var_columns = None
        with open(filename, 'r', encoding=encoding) as f:
            for i, line in enumerate(f.readlines()):
                if i == var_names:
                    names = _return_valid_line(line, separator)
                    break
    else:
        names = var_names
    if var_columns is None:
        var_columns = np.arange(len(names))
    elif len(names) != len(var_columns):
        raise IOError('Length of var_names ({}) must be same as length of var_columns'.format(
            len(names), len(var_columns)))

    # Create data container
    data = {key: [] for key in names}

    # Determine variable units
    if var_units is None:
        _units = [''] * n
    elif isinstance(var_units, int):
        with open(filename, 'r', encoding=encoding) as f:
            for i, line in enumerate(f.readlines()):
                if i == var_units:
                    _units = _return_valid_line(line, separator)
                    break
    else:
        _units = var_units
    # Create unit container
    units = {}
    for i, name in enumerate(names):
        units[name] = _units[i]

    # Now extract the data
    # print('XXX2', filename, len(var_columns), len(names))
    with open(filename, 'r', encoding=encoding) as f:
        for i, line in enumerate(f.readlines()):
            if i >= data_begins_on_row:
                _line = _return_valid_line(line, separator)
                if isinstance(_line, bool) and not _line:
                    continue
                if len(_line) != n:
                    break
                # for j, _item in enumerate(_line):
                #     data[names[j]].append(_item)
                for j in range(len(var_columns)):
                    data[names[j]].append(_line[var_columns[j]])
                    # try:
                    #     data[names[j]].append(_line[var_columns[j]])
                    # except IndexError:
                    #     print('XXX2 ', j, len(_line), names[j], list(data.keys()))

    # TODO
    # Create a 1D xarray.DataArray for each column, with name and unit
    # Create a common xarray.Dataset for all columns

    return data, units


def project_wells_new(filename, working_dir, sheet_name=None):
    """
    Returns a dictionary containing the requested wells from the project excel table
    Avoids using the old "Translate log names" column of the project excel table, and uses
    direct interpretation based on the translate to symbol '->' in the log type columns

    ALSO looks in the sheet named 'Well data' (sheet with data not stored in las format, typically checkshots etc.)
    and adds the data listed there

    :param filename:
        Excel file with well names and filenames
    :param working_dir:
    :param sheet_name:
        str
        Name of the sheet to read well data from.
        Default is 'Well logs', which contains well data from las files only
    :return:
        dict
        dictionary with las file names as keys
    """
    from blixt_utils.utils import isnan
    from blixt_utils.utils import fix_well_name
    from blixt_utils.utils import print_info
    result = {}
    selected_wells = get_selected_wells(filename)
    if sheet_name is None:
        sheet_name = 'Well logs'
    # table = pd.read_excel(filename, header=1, sheet_name=sheet_name, engine='openpyxl')
    xl = pd.ExcelFile(filename)
    table = xl.parse(sheet_name, header=1)
    if 'translate log names' in [xx.lower() for xx in list(table.keys())]:
        error_txt = \
            "Project table contain specific column for log name translation (Translate log name)." + \
            'Old version of project_wells must be used'
        print_info(error_txt, 'error', logger)
        raise IOError(error_txt)
    for i, ans in enumerate(table['Given well name']):
        # skip empty rows
        if not isinstance(ans, str):
            continue
        # skip las files we have chosen to ignore
        if table['Use this file'][i].lower() == 'no':
            continue
        if fix_well_name(ans) in selected_wells:
            temp_dict = {}
            log_dict = {}
            translate_to_string = ''
            for key in list(table.keys()):
                if (key.lower() == 'las file') or (key.lower() == 'use this file'):
                    continue
                elif (key.lower() == 'given well name') or \
                        (key.lower() == 'note'):
                    if isinstance(table[key][i], str):
                        if key.lower() == 'given well name':
                            value = fix_well_name(table[key][i])
                        else:
                            value = table[key][i]
                        temp_dict[key] = value
                    else:
                        temp_dict[key] = None  # avoid NaN
                else:
                    val = table[key][i]
                    if isnan(val):
                        continue
                    else:
                        # First search for the translate to symbol (->) in the log names, and make them available 
                        # for the translator later.
                        this_list = make_clean_list(table[key][i])
                        this_cleaned_list = []
                        for log_name in this_list:
                            if '->' in log_name:
                                translate_to_string += '{}, '.format(log_name)
                                this_cleaned_list.append(log_name.split('->')[0].lower().strip())
                            else:
                                this_cleaned_list.append(log_name.lower())
                        for log_name in this_cleaned_list:
                            log_dict[log_name] = key
                temp_dict['logs'] = log_dict
                if len(translate_to_string) == 0:
                    temp_dict['Translate log names'] = None
                else:
                    temp_dict['Translate log names'] = translate_to_string.rstrip(', ')
            # avoid las file names which aren't correctly given as strings
            if not isinstance(table['las file'][i], str):
                temp_file = False
            else:
                temp_file = test_file_path(table['las file'][i], working_dir)
            if temp_file is False:
                warn_txt = 'Warning, las file {} for well {} does not exist'.format(
                    table['las file'][i], ans)
                print_info(warn_txt, 'warning', logger)
                raise Warning(warn_txt)
            result[test_file_path(table['las file'][i], working_dir)] = temp_dict
        # Test if the sheet 'Well data' exists in the project table, and then add the data files listed there
    if 'Well data' in xl.sheet_names:
        table = xl.parse('Well data', header=1)
        for i, ans in enumerate(table['Given well name']):
            # skip empty rows
            if not isinstance(ans, str):
                continue
            # skip files we have chosen to ignore
            if table['Use this file'][i].lower() == 'no':
                continue
            if fix_well_name(ans) in selected_wells:
                temp_dict = {'logs': {}, 'columns': {}, 'units': {}}
                log_dict = {}
                for key in list(table.keys()):
                    if (key.lower() == 'filename') or (key.lower() == 'use this file'):
                        continue
                    elif key.lower() == 'data begins on line':
                        temp_dict[key] = int(table[key][i]) - 1  # from non-pythonic to pythonic counting!
                    elif ((key.lower() == 'given well name') or (key.lower() == 'note') or
                           (key.lower() == 'separator')):
                        if isinstance(table[key][i], str):
                            if key.lower() == 'given well name':
                                value = fix_well_name(table[key][i])
                            else:
                                value = table[key][i]
                            temp_dict[key] = value
                        elif isinstance(table[key][i], float) or isinstance(table[key][i], int):
                            try:
                                temp_dict[key] = int(table[key][i])
                            except ValueError:
                                temp_dict[key] = None  # avoid NaN
                        else:
                            temp_dict[key] = None  # avoid NaN
                    else:
                        val = table[key][i]
                        if isnan(val):
                            continue
                        else:
                            #  val should be a string of 'name, column_number, unit', where the column number begins at 1
                            segs = val.split(',')
                            log_dict['test'] = val
                            temp_dict['logs'][segs[0]] = key
                            temp_dict['columns'][segs[0]] = int(segs[1]) - 1  # from non-pythonic to pythonic counting!
                            # temp_dict['columns'][segs[0]] = int(segs[1])
                            temp_dict['units'][segs[0]] = segs[2].strip()
                    # temp_dict['logs'] = log_dict
                # avoid file names which aren't correctly given as strings
                if not isinstance(table['filename'][i], str):
                    temp_file = False
                else:
                    temp_file = test_file_path(table['filename'][i], working_dir)
                if temp_file is False:
                    warn_txt = 'Warning, las file {} for well {} does not exist'.format(
                        table['las file'][i], ans)
                    print_info(warn_txt, 'warning', logger)
                    raise Warning(warn_txt)
                result[test_file_path(table['filename'][i], working_dir)] = temp_dict

    return result

                
def project_wells(filename, working_dir):
    """
    Returns a dictionary containing the requested wells from the project excel table

    :param filename:
        Excel file with well names and las files
    :param working_dir:
    :return:
        dict
        dictionary with las file names as keys
    """
    from blixt_utils.utils import fix_well_name
    from blixt_utils.utils import isnan
    from blixt_utils.utils import print_info
    result = {}
    selected_wells = get_selected_wells(filename)
    sheet_name = 'Well logs'
    table = pd.read_excel(filename, header=1, sheet_name=sheet_name, engine='openpyxl')
    for i, ans in enumerate(table['Given well name']):
        # skip empty rows
        if not isinstance(ans, str):
            continue
        # skip las files we have chosen to ignore
        if table['Use this file'][i].lower() == 'no':
            continue
        if fix_well_name(ans) in selected_wells:
            temp_dict = {}
            log_dict = {}
            for key in list(table.keys()):
                if (key.lower() == 'las file') or (key.lower() == 'use this file'):
                    continue
                elif (key.lower() == 'given well name') or \
                        (key.lower() == 'note') or \
                        (key.lower() == 'translate log names'):
                    if isinstance(table[key][i], str):
                        if key.lower() == 'given well name':
                            value = fix_well_name(table[key][i])
                        else:
                            value = table[key][i]
                        temp_dict[key] = value
                    else:
                        temp_dict[key] = None  # avoid NaN
                else:
                    val = table[key][i]
                    if isnan(val):
                        continue
                    else:
                        this_list = make_clean_list(table[key][i], small_cap=True)
                        for log_name in this_list:
                            if '->' in log_name:
                                error_txt = \
                                    "Log name {} contain 'Translate to' symbol, ".format(log_name) + \
                                    'new version of project_wells must be used'
                                print_info(error_txt, 'error', logger)
                                raise IOError(error_txt)
                            log_dict[log_name] = key
                temp_dict['logs'] = log_dict
            # avoid las file names which aren't correctly given as strings
            if not isinstance(table['las file'][i], str):
                temp_file = False
            else:
                temp_file = test_file_path(table['las file'][i], working_dir)
            if temp_file is False:
                warn_txt = 'Warning, las file {} for well {} does not exist'.format(
                    table['las file'][i], ans)
                print_info(warn_txt, 'warning', logger)
                raise Warning(warn_txt)
            result[test_file_path(table['las file'][i], working_dir)] = temp_dict
    return result


def get_selected_wells(filename):
    """
    :param filename
        str
        full path file name of project table excel file
    :return
        list
        list of given well names of selected wells
    """""
    from blixt_utils.utils import fix_well_name
    selected_wells = []
    sheet_name = 'Well settings'
    table = pd.read_excel(filename, header=1, sheet_name=sheet_name, engine='openpyxl')
    for i, ans in enumerate(table['Use']):
        # skip empty rows
        if not isinstance(ans, str):
            continue
        if ans.lower() == 'yes':
            this_well = table['Given well name'][i]
            if this_well not in selected_wells:
                selected_wells.append(fix_well_name(this_well))
    return selected_wells


def invert_well_table(well_table, well_name, rename=True):
    """
    Typically, the "log_table"
    :param well_table:
        dict
        As output from project_wells() above
    :param well_name:
        str
        name of the well we want to extract the "inverted well table" from well_table
    :param rename:
        bool
        if True it uses the "Translate log name" information to rename log names
    :return:
        dict
        As opposed to the commonly used "log_table", which relates a log type with one specific log, this dictionary
        relates a log type with multiple log names
        E.G. {'Resisitivity': ['rdep', 'rmed', 'rsha'], ...}
    """
    out = {}
    rdt = None
    if rename:
        rdt = get_rename_logs_dict(well_table)
    for key in list(well_table.keys()):
        if well_table[key]['Given well name'] == well_name:
            for lname, logtype in well_table[key]['logs'].items():
                _renamed = False
                if logtype not in list(out.keys()):
                    out[logtype] = []
                if rename and (rdt is not None):
                    for to_name, from_names in rdt[key].items():
                        if lname.lower() in from_names:
                            _renamed = True
                            out[logtype].append(to_name.lower())
                if not _renamed:
                    out[logtype].append(lname.lower())
    return out


def get_rename_logs_dict(well_table):
    """
    Interprets the "Translate log names"  keys of the well_table and returns a rename_logs dict.

    :param well_table:
        dict
        as returned from project_wells()
    :return:
        dict or None
        {<las file name> : {'gr': ['gr_cpi'],
                            'neu': ['neu_cpi'],
                            'phit': ['phit_cpi'],
                            ...} }
    """
    rename_logs = {}
    for las_file, val in well_table.items():
        these_rename_logs = {}
        if 'Translate log names' not in list(val.keys()):
            rename_logs[las_file] = None
            continue
        if val['Translate log names'] is None:
            rename_logs[las_file] = None
            continue
        _dict = interpret_rename_string(val['Translate log names'])
        for key in list(_dict.keys()):
            if key in list(these_rename_logs.keys()):
                if not _dict[key] in these_rename_logs[key]:  # only insert same rename pair once
                    these_rename_logs[key].append(_dict[key].lower())
            else:
                these_rename_logs[key] = [_dict[key].lower()]
        if len(these_rename_logs) < 1:
            rename_logs[las_file] = None
        else:
            rename_logs[las_file] = these_rename_logs

    return rename_logs


def project_files_info(filename, sheet_name):
    """
    Interprets a sheet in the project table for information about how to read General ASCII files or if data are
    stored in predefined file formats
    :param filename:
        str
        full path to the project table in excel format
    :param sheet_name:
        str
        Name of the sheet from which the file format description is given
    :return:
        dict
        Dictionary with information about how to read the checkshot or well path files for the wells that belongs to the project
        The dictionary has the following keywords
            :keyword 'filename:
                string
                Full path name of well path / deviation / checkshot file
            :keyword 'file format':
                string
                Name of various well path formats, e.g.:
                    "well trace from petrel"
                    "original survey points"
                    "petrel checkshot"
                    "general ascii" for various formats
            :keyword 'separator':
                string
                'space', 'tab', ',', ';', ...
                String used as separator between data values
                if equal to 'space', any number of spaces is considered a separator
                Not necessary if format is different from 'general ascii'
            :keyword 'data begins on line'
                float or int
                Line number (starting with 1) of where the data section starts
            :keyword 'tvd column':
                int or float
                Column number for TVD values (starting from 1)
                Not necessary if format is different from 'general ascii'
            :keyword 'md column':
                int or float
                Column number for MD values (starting from 1)
                Not necessary if format is different from 'general ascii'
            :keyword 'inclination column':
                int or float
                Column number for inclination values (starting from 1)
                Not necessary if format is different from 'general ascii'

    """
    from blixt_utils.utils import isnan
    result = {}
    table = None
    try:
        table = pd.read_excel(filename, header=1, sheet_name=sheet_name, engine='openpyxl')
    except ValueError:
        raise
    except Exception as e:
        print(e)
    for i, ans in enumerate(table['Use this file']):
        # skip empty rows
        if not isinstance(ans, str):
            continue
        if ans.lower() == 'yes':  # Only use lines in which "Use this file" is set to "Yes"
            temp_dict = {}
            for key in list(table.keys()):
                if (key.lower() == 'use this file') or (key.lower() == 'given well name'):
                    continue
                if isnan(table[key][i]):
                    temp_dict[key.lower()] = None  # avoid NaN
                else:
                    value = table[key][i]
                    temp_dict[key.lower()] = value
            result[table['Given well name'][i].upper()] = temp_dict
    return result


def project_templates(filename):
    from blixt_utils.utils import print_info
    from blixt_utils.utils import isnan
    table = pd.read_excel(filename, header=1, sheet_name='Templates', engine='openpyxl')
    result = {}
    for i, ans in enumerate(table['Log type']):
        if not isinstance(ans, str):
            continue
        result[ans] = {}
        for key in ['bounds', 'center', 'colormap', 'description', 'max', 'min',
                    'scale', 'type', 'unit', 'line color', 'line style', 'line width', 'marker']:
            if key not in list(table.keys()):
                continue
            result[ans][key] = None if isnan(table[key][i]) else table[key][i]
        result[ans]['full_name'] = ans

    # Also add the well settings
    table = pd.read_excel(filename, header=1, sheet_name='Well settings', engine='openpyxl')
    for i, ans in enumerate(table['Given well name']):
        if not isinstance(ans, str):
            continue
        result[ans.upper().strip()] = {}
        for key in ['Color', 'Symbol', 'Content', 'KB', 'UWI', 'UTM', 'X', 'Y', 'Water depth', 'Note']:
            try:
                result[ans.upper().strip()][key.lower()] = None if isnan(table[key][i]) else table[key][i]
            except KeyError as e:
                warn_txt = 'Key {} not found in Well setting sheet in {}'.format(e, filename)
                print_info(warn_txt, 'warning', logger)


    return result


def project_working_intervals(filename):
    table = pd.read_excel(filename, header=4, sheet_name='Working intervals', engine='openpyxl')
    result = {}
    return return_dict_from_tops(table, 'Given well name', 'Interval name', 'Top depth', include_base='Base depth')


def collect_project_wells(well_table, target_dir):
    """
    Copies all las files in the well table (the output from project_wells()) to the folder target_dir.

    :param well_table:
        dictionary, as output from project_wells()
    :return:
    """
    from shutil import copyfile, SameFileError
    for las_file in list(well_table.keys()):
        if os.path.isfile(las_file):
            short_name = os.path.split(las_file)[-1]
            print('Copying file {} to {}'.format(
                short_name, target_dir
            ))
            try:
                copyfile(las_file, os.path.join(target_dir, short_name))
            except SameFileError:
                print('  File {} exists in target directory. Skipping to next'.format(short_name))
                continue


def read_sums_and_averages(filename, header=20):
    table = pd.read_excel(filename, header=header, engine='openpyxl')
    unique_layers = unique_names(table, 'Name', well_names=False)
    answer = {}
    for layer in unique_layers:
        answer[layer] = {}

    for key in list(table.keys()):
        if key == 'Name':
            continue
        for i, value in enumerate(table[key]):
            answer[table['Name'][i]][key] = value

    return answer


def write_sums_and_averages(filename, line_of_data):
    # This function creates xlsx files. so please use Excel to save them as
    # xls files before attempting to load them into RokDoc
    # Additional columns are added after the  'ShaleVolumeAspectRatio' column, which are not
    # read by RokDoc
    if filename.split('.')[-1] == 'xls':
        filename += 'x'

    from openpyxl import load_workbook, Workbook
    if not os.path.isfile(filename):
        print('Creating new RokDoc Sums and Averages file')
        newfile = True
        wb = Workbook()
    else:
        print('Appending to existing RokDoc Sums and averages file')
        newfile = False
        wb = load_workbook(filename)

    ws = wb.active
    if newfile:
        ws.append(['Averages Set output from blixt_rp python library on {}'.format(
            datetime.now().isoformat())])
        ws.append(['Template Version: 1'])
        ws.append(['Depth units:             m'])
        ws.append(['Time units:              ms'])
        ws.append(['Velocity units:          m/s'])
        ws.append(['Density units:           g/cm3'])
        ws.append(['Porosity units:          fract'])
        ws.append(['AI units:                g/cm3_m/s'])
        ws.append(['SI units:                g/cm3_m/s'])
        ws.append(['M units:                 GPa'])
        ws.append(['MU units:                GPa'])
        ws.append(['K (Bulk Modulus) units:  GPa'])
        ws.append(['Lambda units:            GPa'])
        ws.append(['E units:                 GPa'])
        ws.append(['Lambda Mu units:         fract'])
        ws.append(['Mu Rho units:            GPa_g/cm3'])
        ws.append(['Lambda Rho units:        GPa_g/cm3'])
        ws.append(['Saturation units:        fract'])
        ws.append(['Volume units:            fract'])
        ws.append(['TableStart:'])
        ws.append(
            [
                'Name', 'Well', 'ZType', 'TopDepth', 'BaseDepth', 'MidPointDepth',
                'VpMean', 'VsMean', 'RhoMean', 'VpMedian', 'VsMedian', 'RhoMedian',
                'VpMode', 'VsMode', 'RhoMode', 'PorosityType', 'PorosityMean',
                'PorosityStdDev', 'Net', 'NetToGross', 'EpsilonMean', 'DeltaMean',
                'GammaMean', 'EpsilonMedian', 'DeltaMedian', 'GammaMedian',
                'EpsilonMode', 'DeltaMode', 'GammaMode', 'VpStdDev', 'VsStdDev',
                'RhoStdDev', 'EpsilonStdDev', 'DeltaStdDev', 'GammaStdDev',
                'VpVsCorrCoef', 'VpRhoCorrCoef', 'VsRhoCorrCoef', 'AI', 'SI',
                'M', 'MU', 'KBulkModulus', 'PR', 'Lambda', 'E', 'LambdaMu',
                'MuRho', 'LambdaRho', 'ShaleVolumeMean', 'ShaleVolumeStdDev',
                'ShaleVolumeInclusionShape', 'ShaleVolumeAspectRatio', 'Classification',
                'LogsUsed', 'DateAdded'
            ]
        )
    ws.append(line_of_data)
    wb.save(filename)
    wb.close()


def read_pcube_lfc(filename):
    """
    Reads a .lfc file and returns a dictionary with the information the file holds in a format compatible with
    read_sums_and_averages
    :param filename: str
        Full path name of .lfc file
    :return:
        Dictionary with
        { <NAME>:
            {
                'VpMean': XXX,
                'VsMean': XXX,
                'RhoMean': XXX,
                'VpStdDev': XXX,
                'VsStdDev': XXX,
                'RhoStdDev': XXX,
                'VpVsCorrCoef': XXX,
                'VpRhoCorrCoef': XXX,
                'VsRhoCorrCoef': XXX,
                'Classification': XXX,
                'LogsUsed': XXX,
                'DateAdded': XXX
            }
        }
    """
    # create an empty container with the resulting data
    answer = {}
    with open(filename) as f:
        all_lines = f.readlines()
    # First search through the file to extract the name
    name = 'XXX'
    for line in all_lines:
        if 'Name: ' in line:
            name = line.replace('Name: ', '').replace('\n', '')
            answer[name] = {}

    # strings that identifies the data elements of the input and output
    data_ids = [
        ['MeanVp: ', 'VpMean'],
        ['MeanVs: ', 'VsMean'],
        ['MeanRho: ', 'RhoMean'],
        ['StdevVp: ', 'VpStdDev'],
        ['StdevVs: ', 'VsStdDev'],
        ['StdevRho: ', 'RhoStdDev'],
        ['XCCVpVs: ', 'VpVsCorrCoef'],
        ['XCCVpRho: ', 'VpRhoCorrCoef'],
        ['XCCVsRho: ', 'VsRhoCorrCoef'],
        ['# Cutoffs used: ', 'Classification'],
        ['# Logs used: ', 'LogsUsed'],
        ['# Date added: ', 'DateAdded']
    ]

    # loop through file again and search for the data identifiers
    for line in all_lines:
        for data_id in data_ids:
            if data_id[0] in line:
                if '#' in data_id[0]:
                    answer[name][data_id[1]] = line.replace(data_id[0], '').replace('\n', '')
                else:
                    answer[name][data_id[1]] = float(line.replace(data_id[0], '').replace('\n', ''))
    return answer


def write_pcube_lfc(working_dir, results, wi_name, log_table, cutoffs_str, suffix, well_name=None):
    """
    Save the results in a format compatible with the .lfc format used in PCube+ Petrel plug-in from Cegal.
    The units are assumed to be m/s and g/cc
    :return
    """
    from blixt_utils.utils import nan_corrcoef
    from blixt_utils.utils import print_info
    if well_name is not None:
        name = '{} {}{}'.format(well_name, wi_name, suffix)
    else:
        name = '{}{}'.format(wi_name, suffix)
    log_table_str = ''
    for key in log_table:
        log_table_str += '{}: {}, '.format(key, log_table[key])
    log_table_str = log_table_str.rstrip(', ')
    with open(os.path.join(working_dir, '{}.lfc'.format(name)), 'w') as f:
        f.write('Name: {}\n'.format(name))
        f.write('MeanVp: {:.2f}\n'.format(np.nanmean(results[log_table['P velocity'].lower()])))
        f.write('MeanVs: {:.2f}\n'.format(np.nanmean(results[log_table['S velocity'].lower()])))
        f.write('MeanRho: {:.4f}\n'.format(np.nanmean(results[log_table['Density'].lower()])))
        f.write('StdevVp: {:.2f}\n'.format(np.nanstd(results[log_table['P velocity'].lower()])))
        f.write('StdevVs: {:.2f}\n'.format(np.nanstd(results[log_table['S velocity'].lower()])))
        f.write('StdevRho: {:.4f}\n'.format(np.nanstd(results[log_table['Density'].lower()])))
        vp_vs = nan_corrcoef(
            results[log_table['P velocity'].lower()],
            results[log_table['S velocity'].lower()])[0, 1]
        if vp_vs > 0.9:
            warn_txt = 'Correlation between Vp and Vs is spuriously high in {}, replaced with a ccoef of 0.9'.format(name)
            print_info(warn_txt, 'warning', logger)
            vp_vs = 0.9
        f.write('XCCVpVs: {:.2f}\n'.format(vp_vs))
        f.write('XCCVpRho: {:.4f}\n'.format(nan_corrcoef(
            results[log_table['P velocity'].lower()],
            results[log_table['Density'].lower()])[0, 1]))
        f.write('XCCVsRho: {:.4f}\n'.format(nan_corrcoef(
            results[log_table['S velocity'].lower()],
            results[log_table['Density'].lower()])[0, 1]))
        f.write('# Cutoffs used: {}\n'.format(cutoffs_str))
        f.write('# Logs used: {}\n'.format(log_table_str))
        f.write('# Date added: {}\n'.format(datetime.now()))


def read_tops(filename, top=True, zstick='md', frmt=None, only_these_wells=None):
    """

    :param filename:
        str
        full path name to excel sheet with well tops
    :param top:
    :param zstick:
    :param frmt:
        str
        'petrel'
        'npd'
        'rokdoc'
    :param only_these_wells:
        list
        list of well names to look for, so that the reading in can be speeded up
        Populate this list by extracting a well_table, and use the well listed there
        > well_table = rp_utils.io.project_wells(project_table_file)
        > only_these_wells = list(set([x['Given well name'] for x in well_table.values()]))
        NOTE! The naming convention of the wells in the project table file must be the same as the one
        used in tops
    :return:
        dict
        {'well_A name': {'top1 name': top1_depth, 'top2 name': top2_depth, ...},  'well_B name': {...} }
    """
    if frmt == 'petrel':
        return read_petrel_tops(filename, top=top, zstick=zstick, only_these_wells=only_these_wells)
    elif frmt == 'npd':
        return read_npd_tops(filename, top=top, zstick=zstick, only_these_wells=only_these_wells)
    elif frmt == 'rokdoc':
        return read_rokdoc_tops(filename, top=top, zstick=zstick, only_these_wells=only_these_wells)
    else:
        raise IOError('No tops for format {}'.format(frmt))


def read_rokdoc_tops(filename, header=4, top=True, zstick='md', only_these_wells=None):
    """
    :param top:
        bool
        if True, the top of each marker/top is returned
        if False, not implemented
    :param zstick:
        str
        adapted after RokDoc.
        Can be:
            'md', 'tvdkb','twt', 'tvdss',
    :param only_these_wells:
        list
        list of well names to look for, so that the reading in can be speeded up
    """
    if not top:
        raise NotImplementedError('Only top of markers / tops are available')

    if zstick == 'md':
        key_name = 'MD'
    elif zstick == 'tvdkb':
        key_name = 'TVDkb'
    elif zstick == 'twt':
        key_name = 'TWT'
    elif zstick == 'tvdss':
        key_name = 'TVDss'
    else:
        key_name = None
        NotImplementedError('ZSTICK = {} is not implemented'.format(zstick))

    tops = pd.read_excel(filename, header=header, engine='openpyxl')
    return return_dict_from_tops(tops, 'Well Name', 'Horizon', key_name, only_these_wells=only_these_wells)


def read_npd_tops(filename, header=None, top=True, zstick='md', only_these_wells=None):
    """

    :param filename:
    :param header:
    :param top:
    :param zstick:
    :param only_these_wells:
        list
        list of well names to look for, so that the reading in can be speeded up
    :return:
    """
    if zstick != 'md':
        NotImplementedError('ZSTICK = {} is not implemented'.format(zstick))
    if top:
        key_name = 'Top depth [m]'
    else:
        key_name = 'Bottom depth [m]'

    tops = pd.read_excel(filename, engine='openpyxl')
    return return_dict_from_tops(tops, 'Wellbore name', 'Lithostrat. unit', key_name, only_these_wells=only_these_wells)


def read_petrel_tops(filename, header=None, top=True, zstick='md', only_these_wells=None):
    if zstick != 'md':
        NotImplementedError('ZSTICK = {} is not implemented'.format(zstick))
    if not top:
        NotImplementedError('Only MD is implemented for Petrel top files')
        key_name = None
    else:
        key_name = 'MD'

    tops = pd.read_excel(filename, engine='openpyxl')
    return return_dict_from_tops(tops, 'Well identifier', 'Surface', key_name, only_these_wells=only_these_wells)


def write_working_intervals(filename, tops, frmt, well_names=None, interval_names=None, sheet_name=None):
    """
    Writes the tops to the excel file "filename", in the sheet name 'Working intervals'
    If "filename" exists, and is open, it raises a warning

    :param filename:
        str
        full pathname of excel file to write to.
        Assumes we're trying to write to the default project_table.xlsx, in the 'Working intervals' sheet.

    :param tops:
        str or dict
        Either full pathname of file name containing the tops to read from.
        or a dictionary with tops as returned from read_tops()

    :param frmt:
        str
        the file formats supported by blixt_utils.io.read_tops()
        'petrel'
        'npd'
        'rokdoc'
        Not necessary when writing tops that already have been read into "tops"

    :param well_names:
        list
        list of str's
        list of names of the wells we would like to save to to file
        If None, all wells are saved

    :param interval_names:
        list
        list of str's
        list of names of the intervals (working intervals) we would like to save to to file,
        if None, all intervals are saved

    :return:
    """
    if sheet_name is None:
        sheet_name = 'Working intervals'

    # test write access
    taccs = check_if_excelfile_writable(filename)
    if not taccs:
        warn_txt = 'Not possible to write to {}'.format(filename)
        return

    if not os.path.isfile(filename):
        wb = Workbook()
    else:
        wb = load_workbook(filename)

    if sheet_name not in wb.sheetnames:
        print('Creating new sheet')
        ws = wb.create_sheet(sheet_name, -1)
    else:
        print('Opening existing sheet')
        ws = wb[sheet_name]

    # modify first line
    t0 = datetime.now().isoformat()
    if not isinstance(tops, str):
        source_string = 'imported dictionary'
    else:
        source_string = tops
    ws['A1'] = 'Created by {}, at {}, on {} based on tops in {}'.format(
        getpass.getuser(), socket.gethostname(), t0, source_string)

    # test if fifth row exists
    if ws[5][0].value is None:
        ws['A2'] = 'Depth are in meters MD'
        for j, val in enumerate(['Use', 'Given well name', 'Interval name', 'Top depth', 'Base depth', 'Source', 'Note']):
            ws.cell(5, j+1).value = val

    # read in the tops
    if isinstance(tops, str):
        tops = read_tops(tops, frmt=frmt)

    # start appending data
    if well_names is None:
        well_names = list(tops.keys())

    for wname in well_names:
        these_tops = list(tops[wname].keys())
        if len(these_tops) == 0:
            continue  # skip wells without tops
        if interval_names is None:
            int_names = these_tops
            # Add a duplicate of the last interval to avoid running out-of-index
            int_names.append(int_names[-1])
        else:
            int_names = deepcopy(interval_names)

        # Add the 'TD' top name to catch it if it exists
        int_names.append('TD')

        # Find list of common top names
        ct = [tn for tn in int_names if tn in these_tops]
        # Find the index in these_tops to the last common top
        if len(ct) > 0:
            ind = these_tops.index(ct[-1])
            if ind + 1 <= len(these_tops):
                # if this is the last index of these_tops
                ct.append(ct[-1])
            else:
                # Add the next top in these_tops
                ct.append(these_tops[ind+1])
        else:
            ct.append(ct[-1])

        try:
            while 'TD' in ct:
                ct.remove('TD')
        except ValueError as ve:
            print(ve)

        for i in range(len(ct)-1):
            ws.append(['', wname, ct[i], tops[wname][ct[i]], tops[wname][ct[i+1]]])

    wb.save(filename)


def write_regression(filename, reg_params, log_name, well_name, interval_name, reg_type, note=None, sheet_name=None):
    """
    Writes regression (a function optimized to fit some data) to the excel file "filename",
    in the sheet name 'Regressions'
    If "filename" exists, and is open, it raises a warning

    :param filename:
        str
        full pathname of excel file to write to.
        Assumes we're trying to write to the default project_table.xlsx, in the 'Working intervals' sheet.

    :param reg_params:
        list
        A list of parameters, floats, of the regression function.
        E.G. for a linear regression f(t) = A x t + B it will be a two item list [A, B]

    :param log_name:
        str
        Name of the log the regression is calculated on

    :param well_name:
        str
        Name of the well for which the regression is done
        if empty or None, it is valid for all wells

    :param reg_type:
        str
        Type of regression function.
        E.G. for a linear regression, it should be 'Linear'

    :param interval_name:
        str
        Name of the working interval in which the regression is calculated
        if None, all intervals are saved

    :param note:
        str
        Note which describes the regression

    :param sheet_name:
        str
        Name of the sheet the above information will be written to in the filename excel sheet.
        If None, the default "Regressions" work sheet will be used

    :return:
    """
    from blixt_utils.utils import print_info
    allowed_reg_types = ['Linear']
    if reg_type.lower() not in [xx.lower() for xx in allowed_reg_types]:
        warn_txt = 'The given regression type,{}, is not one of the recognized types: {}'.format(
            reg_type, ', '.join(allowed_reg_types)
        )
        print_info(warn_txt, 'warning', logger)
        raise IOError(warn_txt)
    if sheet_name is None:
        sheet_name = 'Regressions'
    if well_name is None or len(well_name) == 0:
        well_name = 'All'
    if interval_name is None or len(interval_name) == 0:
        interval_name = 'All'
    if note is None:
        note = ''

    # test write access
    taccs = check_if_excelfile_writable(filename)
    if not taccs:
        warn_txt = 'Not possible to write to {}'.format(filename)
        return

    if not os.path.isfile(filename):
        wb = Workbook()
    else:
        wb = load_workbook(filename)

    if sheet_name not in wb.sheetnames:
        print('Creating new sheet')
        ws = wb.create_sheet(sheet_name, -1)
    else:
        print('Opening existing sheet')
        ws = wb[sheet_name]

    # test if worksheet contains data
    if ws[2][0].value is None:
        for j, val in enumerate(['Log name', 'Given well name', 'Type', 'Interval name', 'Date', 'Note',
                                 'A', 'B', 'C', 'D', 'E', 'F', 'G']):
            ws.cell(1, j+1).value = val

    # start appending data
    t0 = datetime.now().isoformat()
    ws.append([log_name, well_name, interval_name, reg_type, t0, note, *reg_params])

    wb.save(filename)


def read_regressions(filename, sheet_name=None):
    """
    Reads the excel file 'filename', working sheet 'sheet_name' and returns all regression functions described there
    Args:
        filename:
            str
            Full path name of excel sheet
        sheet_name:
            str
            Name of work sheet which contains the regressions
            Default is 'Regressions'

    Returns:
        dict
        Dictionary with all regressions sorted by log name, well name, interval name and regression type
        Non-unique combinations of log name, well name and interval name are overwritten
    """
    if sheet_name is None:
        sheet_name = 'Regressions'

    table = pd.read_excel(filename, header=1, sheet_name=sheet_name, engine='openpyxl')

    # Create dictionary to contain all regressions
    result = {}
    for i, log_name in enumerate(table['Log name']):
        this_well_name = table['Given well name'][i]
        this_interval_name = table['Interval name'][i]
        this_type = table['Type'][i]

        if log_name not in list(result.keys()):
            result[log_name] = {}
        if this_well_name not in list(result[log_name].keys()):
            result[log_name][this_well_name] = {}
        if this_interval_name not in list(result[log_name][this_well_name].keys()):
            result[log_name][this_well_name][this_interval_name] = {}
        if this_type not in list(result[log_name][this_well_name][this_interval_name].keys()):
            result[log_name][this_well_name][this_interval_name][this_type] = {}

        result[log_name][this_well_name][this_interval_name][this_type]['Date'] = table['Date'][i]
        result[log_name][this_well_name][this_interval_name][this_type]['Note'] = table['Note'][i]
        these_params = []
        for column in ['A', 'B', 'C', 'D', 'F', 'G']:
            if not np.isnan(table[column][i]):
                these_params.append(table[column][i])
        result[log_name][this_well_name][this_interval_name][this_type]['Params'] = these_params

    return result


def read_petrel_points(filename):
    """
    Reads a Petrel points (with attributes) file, and returns a dictionary with a "header" and a "data" key.
    The value belonging to the "data" key is a dictionary with each column of data as a key: value pair
    Args:
        filename:

    Returns:
        dict
    """
    result = {'data': {}}
    keys = []
    key_types = []
    no_strings = False
    data_section = False
    header_section = False
    i = 0
    header_txt = 'Original filename: {}\n'.format(filename)

    if filename is None:
        return None

    xx = ''
    with open(filename, 'r') as f:
        for line in f.readlines():
            if line[0] == '#':
                header_txt += '{}'.format(line)
            if line[:7] == 'BEGIN H':
                header_section = True
                continue
            elif line[:5] == 'END H':
                header_section = False
                data_section = True
                if 'STRING' not in key_types:  # No strings in the data, allows for faster reading
                    no_strings = True
                # add a container for each column of data
                for _x in keys:
                    result['data'][_x] = []
                continue
            if header_section:
                _line = line.split(',')
                if len(_line) == 1:
                    keys.append(_line[0].replace('\n', ''))
                    key_types.append('FLOAT')
                elif len(_line) == 2:
                    keys.append(_line[1].replace('\n', ''))
                    key_types.append(_line[0].replace('\n', ''))
            if data_section:
                if no_strings:
                    _line = line.split()
                    for _i, _key in enumerate(keys):
                        result['data'][_key].append(float(_line[_i]))
                else:
                    # Strings are enclosed in " ", and can contain spaces, so we cant split the line using spaces
                    # Instead we need to first identify all strings that are enclosed in " "
                    match = re.findall("\".*?\"", line)
                    # TODO Now we assume that the number of elements in match is the same as the number of string elements
                    # Replace all the matches above, with any string (without spaces)
                    for _match in match:
                        line = line.replace(_match, 'XXX')
                    _line = line.split()
                    string_counter = 0
                    for _i, _key, _key_type in zip(range(len(keys)), keys, key_types):
                        if 'STRING' in _key_type:
                            result['data'][_key].append(match[string_counter].replace('"', ''))
                            string_counter += 1
                        else:
                            result['data'][_key].append(float(_line[_i]))

    result['header'] = header_txt
    return result


def read_petrel_checkshots(filename, only_these_wells=None):
    # TODO
    # Build in a check to see what the time units are!
    # Now we assume it is in ms
    # If we correct this, then make sure every usage of read_petrel_checkshots is updated
    from blixt_utils.utils import fix_well_name
    checkshots = {}
    keys = []
    data_section = False
    header_section = False
    i = 0
    well_i = None

    if filename is None:
        return None

    with open(filename, 'r') as f:
        for line in f.readlines():
            if line[:7] == 'BEGIN H':
                header_section = True
                continue
            elif line[:5] == 'END H':
                header_section = False
                data_section = True
                continue
            if header_section:
                if line[:4].lower() == 'well':
                    well_i = i
                i += 1
                keys.append(line.strip())
            elif data_section:
                # The well name (within "") may contain spaces, and these spaces must be removed before we
                # can continue
                # First find any string within "" on this line
                match = re.search("\".*?\"", line)
                if match is not None:
                    old_well_name = match.group(0)
                    # Remove any white spaces within the well name
                    new_well_name = ''.join(old_well_name.split())
                    # Then replace the well name with spaces with the new well name
                    line = line.replace(old_well_name, new_well_name)
                data = line.split()
                this_well_name = fix_well_name(data[well_i].replace('"', ''))
                if (only_these_wells is not None) and this_well_name not in only_these_wells:
                    continue
                if this_well_name not in list(checkshots.keys()):
                    checkshots[this_well_name] = {xx: [] for xx in keys}
                else:
                    for j, key in enumerate(keys):
                        checkshots[this_well_name][key].append(my_float(data[j]))

    for this_well_name in list(checkshots.keys()):
        if 'TWT picked' in list(checkshots[this_well_name].keys()):
            checkshots[this_well_name]['TWT'] = checkshots[this_well_name].pop('TWT picked')

    return checkshots


def read_well_headers(filename):
    inside_header = True
    key_replacements = []
    data_column = 1E6
    coord_string = None
    coord_identifier = '# Depth reference system X, Y:'
    column_headers = ['Name', 'Well symbol', 'Surface X', 'Surface Y', 'Well datum value', 'TD (MD)', 'Operator',
                      'TWT auto', 'STRING,Total Depth Type', 'STRING,Well UWI Type', 'STRING,Offshore Block',
                      'STRING,Comment', 'STRING,Well Comment', 'STRING,Offshore Area', 'STRING,Country',
                      'STRING,Well Uwi', 'STRING,ElevationDatum', 'STRING,Location', 'STRING,Well.GroundElevation',
                      'STRING,PathXYOffsetCRS', 'STRING,Elevation', 'STRING,Project.CoordinateSystem']
    for row in open(filename, 'r'):
        if row[0] == '#':
            if row.__contains__(coord_identifier):
                coord_string = row.replace(coord_identifier, '')
            continue
        if row[:10] == 'BEGIN HEAD':
            data_column = 0
        if row[:10] == 'END HEADER':
            inside_header = False
            continue
        if inside_header:
            # Extract the columns we are interested in
            pass

        else: # inside the data section
            # Find all parts of the row that contains a '<key AAAA> BBBB </key>' section, and replace and ignore it
            # TODO handle the replaced xml code too
            match = re.findall("<key (.*?)</key>", row)
            for i, this_match in enumerate(match):
                row = row.replace('<key {}</key>'.format(this_match), 'key_replacement_{}'.format(i))

            # TODO Find any strings specifying a lat / lon (e.g 65°55'1.6900"N) and replace with deg, min, sec
            # match = re.findall('[^ ]*\"[NE]', row)

            # find any string within "" on this line
            match = re.findall("(?:\")(.*?)(?:\")", row)
            # remove any spaces in these strings
            for i, this_match in enumerate(match):
                row = row.replace(this_match, '_'.join(this_match.split()))

            # Now split the line and remove any "
            new_line = [_x.replace('"', '') for _x in row.split()]
            print(len(new_line), new_line[40:])

    print(coord_string)


def test_file_path(file_path, working_dir):
    # Convert backward slashes to forward slashes
    file_path = file_path.replace("\\", "/")
    if os.path.isfile(file_path):
        return file_path
    elif os.path.isfile(os.path.join(working_dir, file_path)):
        return os.path.join(working_dir, file_path)
    else:
        return False


def check_if_excelfile_writable(fnm):
    from openpyxl.utils.exceptions import InvalidFileException
    if os.path.exists(fnm):
        # path exists
        if os.path.isfile(fnm): # is it a file or a dir?
            # also works when file is a link and the target is writable
            # try to open and save it
            try:
                wb = load_workbook(fnm)
            except InvalidFileException as ie:
                print(ie)
                return False
            try:
                wb.save(fnm)
            except PermissionError as pe:
                print(pe)
                print('File is open. Please close it and try again')
                return False
            return True
        else:
            return False # path is a dir, so cannot write as a file

    # target does not exist, check perms on parent dir
    pdir = os.path.dirname(fnm)
    if not pdir:
        pdir = '.'
    # target is creatable if parent dir is writable
    return os.access(pdir, os.W_OK)


def unique_names(table, column_name, well_names=True):
    """
    :param table:
        panda object
        as returned from pandas.read_excel()
    returns the list of unique values in a column named 'column_name'
    """
    from blixt_utils.utils import fix_well_name
    if well_names:
        return [fix_well_name(x) for x in list(set(table[column_name])) if isinstance(x, str)]
    else:
        return [x for x in list(set(table[column_name])) if isinstance(x, str)]


def return_dict_from_excel(file_name, sheet_name, header_line, well_key, depth_key, data_str, depth_is_md=True):
    """
    Reads excel sheet with log data (core samples, cuttings, ...) and returns a dictionary with each well as key, 
    the corresponding item is a dictionary with md (tvd) and data arrays. E.G.
        {'34_3_1S': {
            'depth': np.array(...),
            'toc': np.array(...), 
            'lom': np.array(...),
            ...
            }
        }
    
    Args:
        file_name:
            str
            Full path name of excel file to read
        sheet_name:
            str
            Name of sheet to fetch the data from
        header_line:
            int
            (Pythonic) line number of the header lines (from which the column headers (keys) are taken from)
        well_key:
            str
            Column header from which the well name is taken
        depth_key:
            str
            Column header from which the depth information is taken
        data_str:
            str
            String containing data column header name and data output name in the same format as the 'Translate log names'
            item in the output from project_wells_new()
            e.g. 'TOC.Any->toc, LOM from equivalent Ro->lom'
            Where 'TOC.Any' and 'LOM from equivalent Ro' are the column headers under which the data is found.
            The function interpret_rename_string() must be able to interpret this string
        depth_is_md:
            bool
            If True, the depth information in the excel file is in MD.
            If False, the depth information is interpreted as TVD
        
    Returns:
           dict
            Dictionary with each well as key, the corresponding item is a dictionary with md (tvd) and data arrays.
            E.G.
                {'34_3_1S': {
                    'depth': np.array(...),
                    'toc': np.array(...),
                    'lom': np.array(...),
                    ...
                    }
                }

    """
    from blixt_utils.utils import fix_well_name
    out = {}
    data_keys_dict = interpret_rename_string(data_str, keep_case=True)

    table = pd.read_excel(
        file_name,
        header=header_line,
        sheet_name=sheet_name,
        engine='openpyxl')

    if depth_is_md:
        depth_name = 'depth'
    else:
        depth_name = 'tvd'

    # loop over all rows in the table, and sort the data according to well
    last_well = 'XXX'
    for i, ans in enumerate(table[well_key]):
        this_well = fix_well_name(ans)
        if this_well != last_well:
            # Add new empty data dictionary for this well
            out[this_well] = {key: [] for key in list(data_keys_dict.keys()) + [depth_name]}
        out[this_well][depth_name].append(table[depth_key][i])
        for key in list(data_keys_dict.keys()):
            out[this_well][key].append(table[data_keys_dict[key]][i])
        last_well = this_well

    for well_name in list(out.keys()):
        for key in list(out[well_name]):
            out[well_name][key] = np.array(out[well_name][key])
    return out


def return_dict_from_tops(tops, well_key, top_key, key_name, only_these_wells=None, include_base=None):
    """

    :param tops:
    :param well_key:
        str
        Name of the column which contain the well names
    :param top_key:
        str
        Name of the column which contain the tops / interval names
    :param key_name:
        str
        Name of the column which contain the top depth
    :param only_these_wells:
    :param include_base:
        str
        Name of the column which contain the base depth
        if this is used each top / interval will contain a list of of top and base depth
        else it is just the top depth
    :return:
    """
    from blixt_utils.utils import fix_well_name
    if only_these_wells:
        unique_wells = only_these_wells
    else:
        unique_wells = unique_names(tops, well_key)
    answer = {}

    for well_name in unique_wells:
        answer[well_name] = {}

    for well_name in unique_wells:
        for i, marker_name in enumerate(list(tops[top_key])):
            if fix_well_name(tops[well_key][i]) != well_name:
                continue  # not on the right well
            if include_base is not None:
                answer[well_name][marker_name.upper()] = [tops[key_name][i], tops[include_base][i]]
            else:
                answer[well_name][marker_name.upper()] = tops[key_name][i]

    return answer


def make_clean_list(input_str, small_cap=False, separator=None):
    if not isinstance(input_str, str):
        raise IOError('Only accept strings')
    if separator is None:
        separator = ','
    if small_cap:
        return [x.strip().replace('"', '').lower() for x in input_str.split(separator) if x.strip() != '']
    else:
        return [x.strip().replace('"', '') for x in input_str.split(separator) if x.strip() != '']


def write_las(filename, wh, lh, data, overwrite=False):
    """
    Write the well data to a las format file.

    :param filename:
        str
        full path name to las file
    :param wh:
        core.well.Header
        well header =  w.header, where w is a core.well.Well object
    :param lh:
        core.well.Header
        block header = w.block['Block name'].header, where w is a core.well.Well object
    :param data:
        dict
        data = w.block['Block name'].logs, where w is a core.well.Well object
    :param overwrite:
        bool
        Set to True to allow overwriting an existing las file
    :return:
    """
    from blixt_utils.utils import print_info
    if os.path.isfile(filename) and (not overwrite):
        warn_txt = 'File {} already exist. Write cancelled'.format(filename)
        print_info(warn_txt, 'warning', logger)
        return

    out = (
        '#----------------------------------------------------------------------------\n'
        '~VERSION INFORMATION\n'
        'VERS.            2.0                  :CWLS LOG ASCII STANDARD -VERSION 2.0\n'
        'WRAP.            NO                   :ONE LINE PER DEPTH STEP\n'
        '#\n'
    )

    out += '# {}\n'.format(wh['creation_info'].value)
    if 'note' in list(wh.keys()):
        out += '# NOTE: {}\n'.format(wh['note'].value)
    out += '# Written to las on: {}\n'.format(datetime.now().isoformat())
    out += '# Modified on: {}\n'.format(wh['modification_date'].value)
    for key, value in data.items():
        if value.header['modification_history'] is not None:
            out += '#  Modification: {}: {}\n'.format(key, value.header['modification_history'].replace('\n','\n#   '))

    out += (
        '#--------------------------------------------------------------------\n'
        '~WELL INFORMATION\n'
        '#MNEM .UNIT      DATA                 :DESCRIPTION OF MNEMONIC\n'
        '#----------      ------------         -------------------------------\n'
    )

    # add info about start stop etc. from Block header
    for key in list(lh.keys()):
        if key in ['name', 'creation_info', 'creation_date', 'modification_date', 'well']:
            continue
        out += '{0: <7}.{1: <9}{2: <21}:{3:}\n'.format(
            key.upper(),
            lh[key].unit,
            str(lh[key].value) if lh[key].value is not None else '',
            lh[key].desc.upper()
        )

    # add info well header
    for key in list(wh.keys()):
        if key in ['name', 'note', 'creation_info', 'creation_date', 'modification_date']:
            continue
        out += '{0: <7}.{1: <9}{2: <21}:{3:}\n'.format(
            key.upper(),
            wh[key].unit,
            str(wh[key].value) if wh[key].value is not None else '',
            wh[key].desc.upper()
        )

    out += (
        '#\n'
        '# ----------------------------------------------------------------------------\n'
        '~CURVE INFORMATION\n'
        '# MNEM.UNIT                                         : CURVE DESCRIPTION\n'
        '# ----------                                        -------------------------------\n'
    )
    i = 1
    out += '{0: <20}.{1: <33}: {2: <9}{3:}\n'.format(
        'DEPTH',
        data['depth'].header['unit'],
        i,
        data['depth'].header['desc']
    )
    for key, value in data.items():
        if key == 'depth':
            continue
        i += 1
        out += '{0: <20}.{1: <33}: {2: <9}{3:}\n'.format(
            key.upper(),
            value.header['unit'],
            i,
            value.header['desc']
        )
    out += (
        '#\n'
        '# ----------------------------------------------------------------------------\n'
        '~A                  '
    )

    # write data column headers
    for key in list(data.keys()):
        if key == 'depth':
            continue
        out += '{0: <20}'.format(key.upper())
    out += '\n'

    # start writ11ing data
    for i, md in enumerate(data['depth'].values):
        out += '{0: <20}'.format(md)
        for key in list(data.keys()):
            if key == 'depth':
                continue
            # out += '{0: <20}'.format(
            out += '{:<20.8f}'.format(
                wh['null'].value if np.isnan(data[key].values[i]) else data[key].values[i]
            )
        out += '\n'


    with open(filename, 'w+') as f:
        f.write(out)


def get_las_header(filename, encoding=None):
    """
    Iterates over the las files header lines.

    :param filename:
    :param encoding:
        str
    :return:
    """
    with open(filename, 'r', encoding=encoding) as f:
        for row in f.readlines():
            if '~W' in row:
                break
            else:
                yield row


def get_las_well_info(filename, encoding=None):
    """
    Iterates over the las file well info lines.

    :param filename:
    :param encoding:
        str
    :return:
    """
    well_info_section = False
    curve_info_section = False
    header = True
    with open(filename, 'r', encoding=encoding) as f:
        for row in f.readlines():
            # test wich section you are in
            if '~W' in row:
                well_info_section = True
            if '~C' in row:
                curve_info_section = True

            if curve_info_section:
                break
            if well_info_section:
                header = False
                yield row
            if header:
                continue


def get_las_curve_info(filename, encoding=None):
    """
    Iterates over the las file curve info lines.

    :param filename:
    :param encoding:
        str
    :return:
    """
    curve_info_section = False
    data_section = False
    header = True
    with open(filename, 'r', encoding=encoding) as f:
        for row in f.readlines():
            # test wich section you are in
            if '~C' in row:
                curve_info_section = True
            if '~A' in row:
                data_section = True

            if data_section:
                break
            if curve_info_section:
                header = False
                yield row
            if header:
                continue


def get_las_names_units(filename, encoding=None):
    """
    Iterates over the curve info section of a las file and returns a list of all
    curve names, and a list with their respective units

    :param filename:
    :param encoding:
        str
    :return:
    """
    names = []
    units = []
    for line in get_las_curve_info(filename, encoding=encoding):
        if line[0] in ['~', '#']:
            continue
        # index of seperator
        if re.search("[.]{1}", line) is None:
            print('Caught problem')
            continue
        mnem_end = re.search("[.]{1}", line).end()
        unit_end = mnem_end + re.search("[ ]{1}", line[mnem_end:]).end()
        # divide line
        names.append(line[:mnem_end - 1].strip().lower())
        units.append(line[mnem_end:unit_end].strip())

    return names, units


def get_las_start_data_line(filename, encoding=None):
    """

    :param filename:
    :param encoding:
    :return: Pythonic line number of where the data starts
    """
    with open(filename, 'r', encoding=encoding) as f:
        for i, row in enumerate(f.readlines()):
            if '~A' in row:
                return i + 1
    return None


def well_reader(lines, file_format='las'):
    """
    class handling wells, with logs, and well related information
    The reading .las files is more or less copied from converter.py
        https://pypi.org/project/las-converter/

    """
    from blixt_utils.utils import fix_well_name
    def parse(x):
        try:
            x = int(x)
        except ValueError:
            try:
                x = float(x)
            except ValueError:
                pass
        return x

    def get_current_section(_line):
        if '~V' in _line: return 'version'
        if '~W' in _line: return 'well_info'
        if '~C' in _line: return 'curve'
        if '~P' in _line: return 'parameter'
        if '~O' in _line: return 'other'
        if '~A' in _line: return 'data'
        # ~ Unregistered section
        return None

    def add_section(_well_dict, _section, _mnem, _content):
        if _section == "data":
            if isinstance(_content, list):
                _well_dict[_section][_mnem] = _content
            else:
                _well_dict[_section][_mnem].append(_content)
        elif _section == "other":
            _well_dict[_section] += "".join(str(_content).strip())
        else:
            _well_dict[_section][_mnem] = _content
        return _well_dict

    generated_keys = []
    null_val = None
    section = ""
    length_units = ['m', 'ft']
    accepted_sections = {"version", "well_info", "parameter", "curve"}
    descriptions = []
    curve_names = None
    well_dict = {xx: {} for xx in accepted_sections}
    well_dict["data"] = {}
    well_dict["other"] = ""
    #well_dict = {"version": {}, "well_info": {}, "curve": {}, "parameter": {}, "data": {}, "other": ""}
    if file_format == 'RP well table':
        null_val = 'NaN'

    for line in lines:
        content = {}

        if isinstance(line, bytes):
            line = line.decode("utf-8").strip()

        # line just enter or "\n"
        if len(line) <= 1: continue
        # comment
        if "#" in line: continue

        # section
        if "~" in line:
            section = get_current_section(line)

            # get section version first
            if section == "version":
                continue

            # generate keys of log[data] based on log[curve]
            if section == "data":
                generated_keys = [e.lower() for e in well_dict["curve"].keys()]
                for key in generated_keys:
                    # XXX
                    #key = rename_log_name(key)
                    # inital all key to empty list
                    well_dict = add_section(well_dict, section, key, [])

            continue

        if file_format == 'RP well table':
            if line[:7] == 'Columns':
                section = 'RP header'
                continue  # jump into header

            if section == 'RP header':
                if line[2:5] == ' - ':
                    descriptions.append(line.split(' - ')[-1].strip())

            if line[:7] == 'Well ID':
                # parse curve names
                curve_names = [t.strip().lower() for t in line.split('\t')]
                section = 'dummy_value'

            if line[:4] == '  No':
                # parse line of units
                # unit_names = [t.strip() for t in line.split('\t')]
                unit_names = [t.strip() for t in line.split()]
                unit_names = [t.replace('[', '').replace(']', '') for t in unit_names]

                for this_curve_name, this_unit_name, this_description in zip(curve_names, unit_names, descriptions):
                    well_dict = add_section(well_dict, 'curve',
                                            this_curve_name,
                                            {'api_code': None, 'unit': this_unit_name.lower(), 'desc': this_description}
                                            )
                generated_keys = [key for key in curve_names]
                section = 'data'
                # initiate all key to empty list
                for key in generated_keys:
                    # TODO This will probably fail. Test and compare with 'original' converter
                    well_dict.add_section(well_dict, key, [])
                continue  # jump into data

        # unregistered section
        if section is None: continue

        if section in accepted_sections:
            # index of seperator
            if re.search("[.]{1}", line) is None:
                print('Caught problem')
                continue
            mnem_end = re.search("[.]{1}", line).end()
            unit_end = mnem_end + re.search("[ ]{1}", line[mnem_end:]).end()
            colon_end = unit_end + re.search("[:]{1}", line[unit_end:]).start()

            # divide line
            mnem = line[:mnem_end - 1].strip()
            # XXX
            # mnem = rename_log_name(mnem)
            unit = line[mnem_end:unit_end].strip()
            data = line[unit_end:colon_end].strip()
            # in some las files, the unit is given directly behind the data, e.g "30.0 M"
            # When this is the case, clean the data and try add the unit to the unit
            if (data[-2:].lower().strip() in length_units) or (data[-3:].lower().strip() in length_units):
                _tmp = data.split()
                data = _tmp[0].strip()
                _unit = _tmp[-1].strip().lower()
                if len(unit) == 0:
                    unit = _unit
            desc = line[colon_end + 1:].strip()

            # in some las file, the description contains the column number at the start
            # use a regex to find an initial number, and remove it
            test = re.findall(r"^\d+", desc)
            if len(test) > 0:
                desc = desc.replace(test[0], '')
                desc = desc.strip()

            # convert empty string ("") to None
            if len(data) == 0:
                data = None
            if section == "well_info" and mnem == "NULL":
                # save standard LAS NULL value
                null_val = data.rstrip('0')
                #data = None # this line seems strange and uncessary

            # parse data to type bool or number
            # BUT it also parsed well names as floats, which we should avoid
            if data is not None:
                if desc == 'WELL' or mnem == 'WELL' or mnem == 'UWI':
                    #if section == "well_info" and (mnem == "WELL" or mnem == 'UWI'):
                    # catch well name, harmonize it, and avoid the parse() function
                    data = fix_well_name(data)
                elif data == "NO":
                    data = False
                elif data == "YES":
                    data = True
                else:
                    data = parse(data)

            # dynamic key
            key = "api_code" if section == "curve" else "value"
            content = {
                key: data,
                "unit": unit,
                "desc": desc
            }

            well_dict = add_section(well_dict, section, mnem.lower(), content)

        elif section == "data":
            content = line.split()
            # TODO
            # Some las files contain names enclosed in " ", and these names can contain spaces (see
            # S:\Well\Wells_Norway\QUAD_35\35_4-3\Petrophysics\Interpretation\35_4-3_CPI_OUT_2022-11-06.las )
            # And we need to be smart to separate out these somehow.
            # It is done in not-so-smart way in read_petrel_checkshots():

            for k, v in zip(generated_keys, content):
                #v = float(v) if v != null_val else None
                # replace all null values with np.nan, then we have a unified NaN in all well objects
                try:
                    v = float(v) if v.rstrip('0') != null_val else np.nan
                except ValueError:  # capture weird data
                    print(v, type(v))
                    raise ValueError
                well_dict = add_section(well_dict, section, k.lower(), v)

        elif section == "other":
            well_dict = add_section(well_dict, section, None, line)

    return null_val, generated_keys, well_dict


def well_excel_reader(well_name, file_name, sheet_name, header_line, well_key, depth_key, data_str, depth_is_md=True):
    """

    Args:
        well_name:
            str
            well name
        file_name:
            str
            Full path name of excel file to read
        sheet_name:
            str
            Name of sheet to fetch the data from
        header_line:
            int
            (Pythonic) line number of the header lines (from which the column headers (keys) are taken from)
        well_key:
            str
            Column header from which the well name is taken
        depth_key:
            str
            Column header from which the depth information is taken
        data_str:
            str
            String containing data column header name and data output name in the same format as the 'Translate log names'
            item in the output from project_wells_new()
            e.g. 'TOC.Any->toc, LOM from equivalent Ro->lom'
            Where 'TOC.Any' and 'LOM from equivalent Ro' are the column headers under which the data is found.
            The function interpret_rename_string() must be able to interpret this string
        depth_is_md:
            bool
            If True, the depth information in the excel file is in MD.
            If False, the depth information is interpreted as TVD

    Returns:
        null_val, generated_keys, well_dict
    """
    necessary_input = [
        file_name,
        sheet_name,
        header_line,
        well_key,
        depth_key,
        data_str,
        depth_is_md
    ]
    for inp in necessary_input:
        if inp is None:
            raise IOError('Some mandatory input variables are set to None')

    data_dict = return_dict_from_excel(*necessary_input)

    if well_name not in list(data_dict.keys()):
        raise IOError('Current well {} is not listed in file {}'.format(
            well_name, os.path.basename(file_name)
        ))

    null_val = 'NaN'

    generated_keys = [xx.lower() for xx in list(data_dict[well_name].keys())]
    accepted_sections = {"version", "well_info", "parameter", "curve", "data"}
    well_dict = {xx: {} for xx in accepted_sections}
    well_dict['version'] = {'vers': {'value': 'xlsx', 'unit': '', 'desc': 'Read from excel table of well data'}}
    # TODO
    # Make it possible to add units and description to data read from Excel files too
    for key in generated_keys:
        well_dict['curve'][key] = {'api_code': None, 'unit': None, 'desc': None}
        well_dict['data'][key] = data_dict[well_name][key]

    well_dict['well_info']['strt'] = {'value': np.nanmin(well_dict['data']['depth']), 'unit': '', 'desc': ''}
    well_dict['well_info']['stop'] = {'value': np.nanmax(well_dict['data']['depth']), 'unit': '', 'desc': ''}
    well_dict['well_info']['step'] = {'value': None, 'unit': '', 'desc': ''}
    well_dict['well_info']['well'] = {'value': well_name, 'unit': '', 'desc': 'WELL'}

    return null_val, generated_keys, well_dict


def interpret_rename_string(rename_string, keep_case=False):
    """
    creates a rename dictionary ({'VCL': 'VSH', 'Vp': 'Vp_dry'}) from input string
    :param rename_string:
        str
        renaming defined by "VSH->VCL, Vp_dry->Vp"
    :param keep_case:
        bool
        If true, the letter case is unchanged.
        Default False, which changes all cases to lower
    :return:
        dict or None
    """
    from blixt_utils.utils import print_info
    if len(rename_string) < 3:
        return None

    return_dict = {}
    for pair in rename_string.split(','):
        if '->' not in pair:
            continue
        names = pair.split('->')
        if len(names) > 2:
            warn_txt = "Translation pairs should be separated by ',': ".format(pair)
            print_info(warn_txt, 'warning', logger)
            continue
        if keep_case:
            return_dict[names[1].strip()] = names[0].strip()
        else:
            return_dict[names[1].strip().lower()] = names[0].strip().lower()
    if len(return_dict) < 1:
        return None
    else:
        return return_dict


def interpret_cutoffs_string(cutoffs_string):
    """
    creates a cutoffs dictionary ({'VCL': ['<', 0.5], 'PHIE': ['>', 0.1]}) from input string
    :param cutoffs_string:
        str
        renaming defined by "VCL>0.8, PHIE>0.1"
    :return:
        dict or None
    """
    from blixt_utils.utils import print_info
    if len(cutoffs_string) < 3:
        return None

    return_dict = {}
    for pair in cutoffs_string.split(','):
        m_symb = None
        # search for masking symbol
        if '==' in pair:
            m_symb = '=='  # equal
        elif '<=' in pair:
            m_symb = '<='  # less or equal
        elif '>=' in pair:
            m_symb = '>='  # greater or equal
        elif '!=' in pair:
            m_symb = '!='  # not equal
        elif '>' in pair:
            m_symb = '>'  # greater
        elif '<' in pair:
            m_symb = '<'  # less
        else:
            warn_txt = 'No valid masking symbol given in {}'.format(pair)
            print_info(warn_txt, 'warning', logger)
            continue

        tmp = pair.split(m_symb)
        if len(tmp) > 2:
            warn_txt = "Something fishy in cutoffs string: ".format(cutoffs_string)
            print_info(warn_txt, 'warning', logger)
            continue
        return_dict[tmp[0].strip().lower()] = [m_symb, float(tmp[1].strip())]
    if len(return_dict) < 1:
        return None
    else:
        return return_dict


def my_float(string):
    try:
        return float(string)
    except ValueError:
        try:
            return float(string.replace(',', '.'))
        except ValueError:
            return string


def _split(_string, _separator):
    """

    :param _string:
    :param _separator:
        str
        'space', 'tab', ',', ';', ...
    :return:
    """
    if 'space' in _separator.lower():
        _this_list = _string.split()
    elif 'tab' in _separator.lower():
        _this_list = _string.split('\t')
    else:
        _this_list = _string.split(_separator)
    clean_list = []
    for item in _this_list:
        if item == '':
            continue
        # Only convert pure numbers, and decimal numbers, to float
        if item.strip().replace('.', '', 1).replace('-', '', 1).isdigit():
            clean_list.append(float(item.strip()))
        else:
            clean_list.append(my_float(item.strip().replace('"', '')))
    return clean_list


def _return_valid_line(
        _string: str,
        _separator: str,
):
    """
    Splits up the string into data items, which can be both floats and strings.
    For empty lines it return False

    :param _string:
    :param _separator:
    :return:
        list or False
    """
    # Skip empty lines!
    if len(_string.strip().replace('\t', '')) == 0:
        return False

    # Strings are enclosed in " ", and can contain spaces, so we cant split the line using spaces
    # Instead we need to first identify all strings that are enclosed in " "
    match = re.findall("\".*?\"", _string)
    if len(match) > 0:
        # Replace all the spaces within the matches above with any string (without spaces)
        for _match in match:
            if len(_match) > 2:
                _string = _string.replace(_match, _match.replace(' ', 'XXX'))

    _data = _split(_string, _separator)

    if len(match) > 0:
        # Now replace the above XXX with space again
        for i, _item in enumerate(_data):
            if isinstance(_item, str):
                _data[i] = _item.replace('XXX', ' ')

    return _data


def filetype(filename):
    if os.path.isdir(filename):
        return 'folder'
    else:
        return os.path.splitext(filename)[-1].lower().replace('.', '')

def example_parsing_webpage():
    """
    npd beautifulsoup parsing
    :return:
    """
    from bs4 import BeautifulSoup
    import requests

    #r = requests.get("https://factpages.sodir.no/en/wellbore/PageView/Exploration/All/9621")
    r = requests.get("https://factpages.sodir.no/en/wellbore/PageView/Exploration/All/9401")
    # soup = BeautifulSoup(r.content, 'html.parser')
    soup = BeautifulSoup(r.content, 'html5lib')
    table = soup.find('table', attrs={'class': 'a1448 general-info-table uk-table-striped'})
    for row in table.findAll('td', attrs={'class': 'a1256c r6'}):
    # for row in table.findAll('td', attrs={'class': 'a1269c r6'}):
    # for row in table.findAll('td', attrs={'class': 'a850c r6'}):
        print(row.text)

    for row in table.findAll('td', attrs={'class': 'a1269c r6'}):
        print(row.text)


if __name__ == '__main__':
    example_parsing_webpage()
